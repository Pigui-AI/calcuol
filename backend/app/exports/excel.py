"""Exportación a Excel — sección 13.1. Cada workbook referencia el run que lo originó."""
import os
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import SimulationRun, MonthlyProjection, Client, Project, ArInvoice
from app.engine.money import D

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
MONEY_FMT = "#,##0.00"
COUNT_FMT = "#,##0.0"

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports_files")

PNL_ROWS = [
    ("Ingresos por comisiones", "rev.commission"),
    ("Ingresos por suscripciones", "rev.subscriptions"),
    ("Ingresos por IA/tokens", "rev.tokens"),
    ("Ingresos totales", "pnl.revenue"),
    ("Costos variables", "pnl.variable_costs"),
    ("Margen bruto", "pnl.gross_margin"),
    ("Contribution Margin", "pnl.contribution_margin"),
    ("OPEX (fijos + adquisición)", "pnl.opex"),
    ("Costos de campañas", "cost.campaigns"),
    ("Fees de procesamiento", "cost.processing_fees"),
    ("Nómina (hiring plan)", "cost.hiring"),
    ("EBITDA", "pnl.ebitda"),
]

CASH_ROWS = [
    ("Cobro inmediato (Stripe/Pigui Scan)", "cash.collected_immediate"),
    ("Cobranza de AR", "ar.collections"),
    ("Entradas totales", "cash.in_total"),
    ("Salidas totales", "cash.out_total"),
    ("Pago de redenciones de puntos", "cash.points_paid_out"),
    ("Liquidaciones a negocios", "stl.paid_out"),
    ("Payable a negocios (saldo)", "stl.payable_end"),
    ("Flujo neto", "cash.net"),
    ("Caja al cierre", "cash.balance_end"),
    ("AR nueva", "ar.new"),
    ("AR saldo", "ar.balance_end"),
    ("Incobrables", "ar.writeoffs"),
]

PLAN_ROWS = [
    ("Clientes B2B al cierre", "b2b.clients_end", COUNT_FMT),
    ("Altas activadas", "b2b.adds_activated", COUNT_FMT),
    ("Capacidad de onboarding", "b2b.onboarding_capacity", COUNT_FMT),
    ("Churn B2B", "b2b.churned", COUNT_FMT),
    ("Consumidores activos", "b2c.consumers_end", COUNT_FMT),
    ("Compradores", "b2c.buyers", COUNT_FMT),
    ("Transacciones", "b2c.transactions", COUNT_FMT),
    ("GMV", "tx.gmv", MONEY_FMT),
    ("Utilidad elegible", "tx.eligible_utility", MONEY_FMT),
    ("Comisión Pigui", "rev.commission", MONEY_FMT),
    ("Suscriptores activos", "subs.active_total", COUNT_FMT),
    ("Trials iniciados", "subs.trial_starts", COUNT_FMT),
    ("MRR al cierre", "rev.mrr.end", MONEY_FMT),
    ("Tokens consumidos", "tokens.units.consumed", COUNT_FMT),
    ("Tokens overage", "tokens.units.overage", COUNT_FMT),
    ("Puntos emitidos (valor)", "points.emitted", MONEY_FMT),
    ("Utilidad neta del negocio", "tx.business_net", MONEY_FMT),
    ("Saldo de puntos (pasivo)", "points.balance_end", MONEY_FMT),
    ("Headcount", "hiring.headcount", COUNT_FMT),
    ("Nómina (hiring plan)", "cost.hiring", MONEY_FMT),
    ("Ingresos Pigui", "pnl.revenue", MONEY_FMT),
    ("EBITDA", "pnl.ebitda", MONEY_FMT),
    ("Caja", "cash.balance_end", MONEY_FMT),
]

CAMPAIGN_ROWS = [
    ("Campañas activas", "camp.active_count", COUNT_FMT),
    ("Puntos extra de campañas", "camp.extra_points", COUNT_FMT),
    ("GMV incremental", "camp.gmv_incremental", MONEY_FMT),
    ("Ingreso incremental", "camp.revenue_incremental", MONEY_FMT),
    ("Costos de campañas", "cost.campaigns", MONEY_FMT),
    ("ROI de campañas", "camp.roi", "0.00"),
    ("Puntos emitidos (valor)", "points.emitted", MONEY_FMT),
    ("Intentos de redención", "points.funnel.intents", COUNT_FMT),
    ("Puntos redimidos", "points.redeemed", MONEY_FMT),
    ("Puntos expirados", "points.expired", MONEY_FMT),
    ("Saldo de puntos (pasivo)", "points.balance_end", MONEY_FMT),
]

SETTLEMENT_ROWS = [
    ("Transacciones vía Stripe", "tx.count_stripe", COUNT_FMT),
    ("Transacciones en caja", "tx.count_cash", COUNT_FMT),
    ("GMV vía Stripe", "tx.gmv_stripe", MONEY_FMT),
    ("GMV en caja", "tx.gmv_cash", MONEY_FMT),
    ("Cobro bruto vía Stripe (settlements)", "stl.gross_collected", MONEY_FMT),
    ("Fees de procesamiento", "cost.processing_fees", MONEY_FMT),
    ("Neto a liquidar a negocios", "stl.merchant_due", MONEY_FMT),
    ("Liquidaciones a negocios", "stl.paid_out", MONEY_FMT),
    ("Payable a negocios (saldo)", "stl.payable_end", MONEY_FMT),
]

CAMPAIGN_SUMMARY_ROWS = [
    ("Gasto total en campañas", "total_spend", MONEY_FMT),
    ("Puntos extra totales", "total_extra_points", COUNT_FMT),
    ("GMV incremental total", "total_gmv_incremental", MONEY_FMT),
    ("Ingreso incremental total", "total_revenue_incremental", MONEY_FMT),
    ("ROI total", "roi_total", "0.00"),
]

UE_ROWS = [
    ("ARPA (ingreso por cliente)", "ue.arpa", MONEY_FMT),
    ("Contribution margin por cliente", "ue.cm_per_client", MONEY_FMT),
    ("CAC", "ue.cac", MONEY_FMT),
    ("LTV", "ue.ltv", MONEY_FMT),
    ("LTV/CAC", "ue.ltv_cac", "0.00"),
    ("Payback (meses)", "ue.payback_months", "0.0"),
    ("Take rate sobre GMV", "ue.take_rate", "0.00%"),
    ("Burn neto", "kpi.burn_net", MONEY_FMT),
    ("Runway (meses)", "kpi.runway_months", "0.0"),
]

# --- Fase 6: suscripciones detalladas, tokens/IA y hiring plan ---
# Las métricas existen siempre en runs 1.3.0 (cero con los motores apagados);
# en runs previos faltan y las celdas quedan vacías sin romper el export.

SUBSCRIPTION_ROWS = [
    ("MRR inicial", "rev.mrr.start", MONEY_FMT),
    ("MRR nuevo", "rev.mrr.new", MONEY_FMT),
    ("MRR expansión", "rev.mrr.expansion", MONEY_FMT),
    ("MRR contracción", "rev.mrr.contraction", MONEY_FMT),
    ("MRR perdido (churn)", "rev.mrr.churned", MONEY_FMT),
    ("MRR al cierre", "rev.mrr.end", MONEY_FMT),
    ("Trials iniciados", "subs.trial_starts", COUNT_FMT),
    ("Conversiones de trial", "subs.conversions", COUNT_FMT),
    ("Suscriptores activos", "subs.active_total", COUNT_FMT),
]

TOKEN_ROWS = [
    ("Tokens consumidos", "tokens.units.consumed", COUNT_FMT),
    ("Tokens incluidos", "tokens.units.included", COUNT_FMT),
    ("Tokens incluidos expirados", "tokens.units.included_expired", COUNT_FMT),
    ("Crédito inicial usado", "tokens.units.credit_used", COUNT_FMT),
    ("Crédito inicial expirado", "tokens.units.credit_expired", COUNT_FMT),
    ("Tokens overage", "tokens.units.overage", COUNT_FMT),
    ("Tokens de recarga", "tokens.units.recharge", COUNT_FMT),
    ("Ingresos por IA/tokens", "rev.tokens", MONEY_FMT),
    ("Ingresos por overage", "rev.tokens.overage", MONEY_FMT),
    ("Ingresos por recargas", "rev.tokens.recharges", MONEY_FMT),
    ("Costo de tokens (proveedor)", "cost.tokens", MONEY_FMT),
    ("Margen unitario", "tokens.unit_margin", MONEY_FMT),
    ("Margen % de tokens", "tokens.margin_pct", "0.00%"),
]

HIRING_ROWS = [
    ("Headcount", "hiring.headcount", COUNT_FMT),
    ("Nómina (hiring plan)", "cost.hiring", MONEY_FMT),
    ("Capacidad de onboarding", "b2b.onboarding_capacity", COUNT_FMT),
    ("Altas activadas", "b2b.adds_activated", COUNT_FMT),
]

TRIAL_KIND_LABELS = {
    "none": "Sin trial",
    "sin_tarjeta_15": "Sin tarjeta (15 días)",
    "con_tarjeta_30": "Con tarjeta (30 días)",
}


def _num(value):
    """Convierte valores string del snapshot/logs a número para la celda.
    None o vacío ⇒ celda vacía; si no es numérico, se vuelca tal cual."""
    if value in (None, ""):
        return None
    try:
        return float(D(value))
    except Exception:
        return value


def _month_of(value, months):
    """Índice de mes (1-based) → etiqueta 'YYYY-MM'; fuera de rango o no
    numérico se vuelca tal cual (defensivo con logs de runs viejos)."""
    try:
        idx = int(value)
    except (TypeError, ValueError):
        return value
    if 1 <= idx <= len(months):
        return months[idx - 1]
    return idx


def _style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _metric_matrix(db: Session, run_id: str) -> dict:
    rows = db.execute(select(MonthlyProjection).where(MonthlyProjection.run_id == run_id))
    metrics: dict[str, dict[int, Decimal]] = {}
    for r in rows.scalars():
        metrics.setdefault(r.metric_key, {})[r.month_index] = D(r.value)
    return metrics


def _write_metric_sheet(ws, title, months, metrics, rows_def, currency):
    ws.append([title] )
    ws["A1"].font = TITLE_FONT
    ws.append([f"Moneda: {currency}. Valores mensuales."])
    ws.append([])
    header = ["Concepto"] + months
    ws.append(header)
    _style_header(ws, 4, len(header))
    for defn in rows_def:
        label, key = defn[0], defn[1]
        fmt = defn[2] if len(defn) > 2 else MONEY_FMT
        series = metrics.get(key, {})
        row = [label] + [series.get(i + 1) for i in range(len(months))]
        ws.append(row)
        for col in range(2, len(months) + 2):
            ws.cell(row=ws.max_row, column=col).number_format = fmt
    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 34
    for i in range(len(months)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12


def _sheet_run_header(ws, run: SimulationRun, title: str):
    """Cabecera con referencia al run y sus hashes (mismo patrón del README)."""
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.append(["Run de simulación", run.id])
    ws.append(["Hash de inputs", run.input_hash])
    ws.append(["Hash de outputs", run.output_hash or ""])
    ws.append(["Versión del motor", run.engine_version])
    ws.append([])


def _append_metric_matrix(ws, months, metrics, rows_def):
    """Matriz mensual de métricas anexada en la posición actual de la hoja."""
    header = ["Concepto"] + months
    ws.append(header)
    _style_header(ws, ws.max_row, len(header))
    for label, key, fmt in rows_def:
        series = metrics.get(key, {})
        ws.append([label] + [series.get(i + 1) for i in range(len(months))])
        for col in range(2, len(months) + 2):
            ws.cell(row=ws.max_row, column=col).number_format = fmt
    ws.column_dimensions["A"].width = 34
    for i in range(len(months)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12


def generate_workbook(db: Session, run: SimulationRun) -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    snapshot = run.snapshot
    project_info = snapshot["project"]
    scenario_info = snapshot["scenario"]
    currency = project_info["base_currency"]
    summary = (run.logs or {}).get("summary", {})
    metrics = _metric_matrix(db, run.id)
    from app.engine.simulator import month_label
    months = [month_label(project_info["start_month"], i)
              for i in range(1, run.horizon_months + 1)]

    wb = Workbook()

    # --- README ---
    ws = wb.active
    ws.title = "README"
    meta = [
        ("Pigui Financial Engine — Export", ""),
        ("Proyecto", project_info["name"]),
        ("Escenario", f"{scenario_info['name']} ({scenario_info['type']})"),
        ("Run de simulación", run.id),
        ("Hash de inputs", run.input_hash),
        ("Hash de outputs", run.output_hash or ""),
        ("Versión del motor", run.engine_version),
        ("Horizonte (meses)", run.horizon_months),
        ("Mes de inicio", project_info["start_month"]),
        ("Moneda", currency),
        ("Generado (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Advertencia", "Proyección basada en supuestos e hipótesis; no es contabilidad oficial."),
        ("Punto de equilibrio", summary.get("breakeven_label") or "No alcanzado en el horizonte"),
        ("Necesidad de capital", summary.get("funding_need", "")),
        ("Caja mínima", summary.get("min_cash", "")),
    ]
    for label, value in meta:
        ws.append([label, value])
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60

    # --- Assumptions ---
    ws = wb.create_sheet("Assumptions")
    ws.append(["Supuesto", "Valor", "Origen", "Tipo de dato"])
    _style_header(ws, 1, 4)
    origins = snapshot.get("assumption_origins", {})
    for key, value in sorted(snapshot["assumptions"].items()):
        ws.append([key, value, origins.get(key, ""), "hipótesis/declarado"])
    derived = summary.get("derived_inputs", {})
    if derived:
        ws.append([])
        ws.append(["Valores derivados del portafolio (estimado)", "", "", ""])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        for key, info in derived.items():
            ws.append([key, info.get("to"), info.get("source"), "estimado"])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24

    # --- Clients ---
    ws = wb.create_sheet("Clients")
    ws.append(["Cliente", "Industria", "Estado", "Sucursales", "Ventas base/mes",
               "Transacciones/mes", "Ticket", "Margen %", "Consumidores activos"])
    _style_header(ws, 1, 9)
    clients = db.execute(select(Client).where(Client.project_id == project_info["id"],
                                              Client.status != "archived")).scalars().all()
    for c in clients:
        b = c.baseline
        branches = sum(len(brand.branches) for brand in c.brands)
        ws.append([
            c.trade_name, c.industry, c.status, branches,
            float(b.avg_monthly_sales) if b else None,
            float(b.avg_monthly_transactions) if b else None,
            float(b.avg_ticket) if b else None,
            float(b.margin_pct) if b else None,
            b.active_consumers if b else None,
        ])
    for col, width in (("A", 28), ("B", 16), ("E", 16), ("F", 16)):
        ws.column_dimensions[col].width = width

    # --- Catalog ---
    ws = wb.create_sheet("Catalog")
    ws.append(["Cliente", "Sucursal", "Tipo", "Artículo", "SKU", "Precio", "Costo directo",
               "Margen", "Margen %", "Elegible rewards"])
    _style_header(ws, 1, 10)
    for c in clients:
        for brand in c.brands:
            for branch in brand.branches:
                for item in branch.catalog_items:
                    price, cost = D(item.sale_price), D(item.direct_cost)
                    margin = price - cost
                    ws.append([c.trade_name, branch.name, item.type, item.name, item.sku,
                               float(price), float(cost), float(margin),
                               float(margin / price) if price > 0 else None,
                               "Sí" if item.reward_eligible else "No"])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["D"].width = 30

    # --- Monthly Plan / P&L / Cash Flow / Unit Economics ---
    _write_metric_sheet(wb.create_sheet("Monthly Plan"), "Business plan mensual",
                        months, metrics, PLAN_ROWS, currency)
    _write_metric_sheet(wb.create_sheet("P&L"), "Estado de resultados",
                        months, metrics, PNL_ROWS, currency)
    _write_metric_sheet(wb.create_sheet("Cash Flow"), "Flujo de efectivo",
                        months, metrics, CASH_ROWS, currency)
    _write_metric_sheet(wb.create_sheet("Unit Economics"), "Unit economics",
                        months, metrics, UE_ROWS, currency)

    # --- Scenarios (resumen anual del run) ---
    ws = wb.create_sheet("Scenarios")
    ws.append(["Resumen anual", scenario_info["name"]])
    ws["A1"].font = TITLE_FONT
    ws.append([])
    ws.append(["Año", "Ingresos", "GMV", "EBITDA", "OPEX", "Clientes cierre",
               "Consumidores cierre", "Caja cierre"])
    _style_header(ws, 3, 8)
    for row in summary.get("annual", []):
        ws.append([row["year"], float(row["revenue"]), float(row["gmv"]),
                   float(row["ebitda"]), float(row["opex"]),
                   float(row["clients_end"]) if row["clients_end"] else None,
                   float(row["consumers_end"]) if row["consumers_end"] else None,
                   float(row["cash_end"]) if row["cash_end"] else None])
        for col in range(2, 9):
            ws.cell(row=ws.max_row, column=col).number_format = MONEY_FMT
    for col in ("B", "C", "D", "E", "H"):
        ws.column_dimensions[col].width = 15

    # --- Campaigns & Points (fase 5; métricas siempre emitidas, en cero si el motor está apagado) ---
    ws = wb.create_sheet("Campaigns & Points")
    _sheet_run_header(ws, run, "Campañas y puntos")
    ws.append(["Campañas congeladas en el snapshot"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    campaigns = snapshot.get("campaigns", []) or []
    if campaigns:
        ws.append(["Campaña", "Tipo", "Ventana", "Efecto", "Valor"])
        _style_header(ws, ws.max_row, 5)
        for camp in campaigns:
            window = f"Meses {camp.get('start_month', '')}–{camp.get('end_month', '')}"
            effects = camp.get("effects", {}) or {}
            first = True
            for key in sorted(effects):
                ws.append([camp.get("name", "") if first else "",
                           camp.get("campaign_type", "") if first else "",
                           window if first else "", key, effects[key]])
                first = False
            if first:  # campaña sin efectos registrados
                ws.append([camp.get("name", ""), camp.get("campaign_type", ""), window, "", ""])
    else:
        ws.append(["Sin campañas congeladas en este run"])
    ws.append([])
    _append_metric_matrix(ws, months, metrics, CAMPAIGN_ROWS)
    ws.append([])
    ws.append(["Resumen de campañas del horizonte"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    camp_summary = summary.get("campaigns", {}) or {}
    for label, key, fmt in CAMPAIGN_SUMMARY_ROWS:
        value = camp_summary.get(key)
        ws.append([label, float(D(value)) if value not in (None, "") else None])
        ws.cell(row=ws.max_row, column=2).number_format = fmt

    # --- Settlements & AR (fase 5) ---
    ws = wb.create_sheet("Settlements & AR")
    _sheet_run_header(ws, run, "Liquidaciones a negocios y cuentas por cobrar")
    _append_metric_matrix(ws, months, metrics, SETTLEMENT_ROWS)
    ws.append([])
    ws.append(["Facturas de AR del run (ruta en caja)"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    ws.append(["Número", "Mes de emisión", "Monto", "Mes de vencimiento",
               "Cobro esperado", "Castigo esperado", "Estado"])
    _style_header(ws, ws.max_row, 7)
    invoices = db.execute(select(ArInvoice).where(ArInvoice.run_id == run.id)
                          .order_by(ArInvoice.month_index)).scalars().all()
    if invoices:
        for inv in invoices:
            ws.append([inv.invoice_number, inv.month_label, float(D(inv.amount)),
                       inv.due_month_label, float(D(inv.expected_collection)),
                       float(D(inv.expected_writeoff)), inv.status])
            for col in (3, 5, 6):
                ws.cell(row=ws.max_row, column=col).number_format = MONEY_FMT
    else:
        ws.append(["Sin facturas de AR para este run"])

    logs = run.logs or {}

    # --- Subscriptions (fase 6; pantallas 45/47/62) ---
    ws = wb.create_sheet("Subscriptions")
    _sheet_run_header(ws, run, "Suscripciones, trials y MRR bridge")
    ws.append(["Planes de suscripción congelados en el snapshot"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    plans = snapshot.get("subscription_plans", []) or []
    if plans:
        plan_names = {p.get("id"): p.get("name", "") for p in plans}
        ws.append(["Plan", "Precio mensual", "Moneda", "Trial", "Conversión de trial",
                   "Adopción", "Mes de inicio", "Ramp (meses)", "Churn mensual",
                   "Upgrade a plan", "Créditos de tokens incluidos"])
        _style_header(ws, ws.max_row, 11)
        for p in plans:
            upgrade_id = p.get("upgrade_to_plan_id")
            trial_kind = p.get("trial_kind", "")
            ws.append([
                p.get("name", ""),
                _num(p.get("price_monthly")),
                p.get("currency", ""),
                TRIAL_KIND_LABELS.get(trial_kind, trial_kind),
                _num(p.get("trial_conversion")),
                _num(p.get("adoption_rate")),
                p.get("start_month"),
                p.get("ramp_months"),
                _num(p.get("churn_rate")),
                plan_names.get(upgrade_id, upgrade_id or ""),
                _num(p.get("included_token_credits")),
            ])
            row = ws.max_row
            ws.cell(row=row, column=2).number_format = MONEY_FMT
            for col in (5, 6, 9):
                ws.cell(row=row, column=col).number_format = "0.00%"
            ws.cell(row=row, column=11).number_format = COUNT_FMT
    else:
        ws.append(["Sin planes de suscripción congelados en este run"])
    ws.append([])
    _append_metric_matrix(ws, months, metrics, SUBSCRIPTION_ROWS)
    ws.append([])
    ws.append(["Cohortes de trial del run (conversión por tipo)"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    cohorts = logs.get("subs_cohorts", []) or []
    if cohorts:
        ws.append(["Plan", "Tipo de trial", "Mes de cohorte", "Inicios",
                   "Mes de decisión", "Conversiones", "Tasa de conversión"])
        _style_header(ws, ws.max_row, 7)
        for c in cohorts:
            trial_kind = c.get("trial_kind", "")
            ws.append([
                c.get("plan_name", "") or c.get("plan_id", ""),
                TRIAL_KIND_LABELS.get(trial_kind, trial_kind),
                _month_of(c.get("cohort_month"), months),
                _num(c.get("starts")),
                _month_of(c.get("decision_month"), months),
                _num(c.get("conversions")),
                _num(c.get("conversion_rate")),
            ])
            row = ws.max_row
            for col in (4, 6):
                ws.cell(row=row, column=col).number_format = COUNT_FMT
            ws.cell(row=row, column=7).number_format = "0.00%"
    else:
        ws.append(["Sin cohortes de trial en este run"])

    # --- Tokens (fase 6; pantallas 46/63) ---
    ws = wb.create_sheet("Tokens")
    _sheet_run_header(ws, run, "IA y tokens — unidades, dinero y margen")
    _append_metric_matrix(ws, months, metrics, TOKEN_ROWS)
    ws.append([])
    ws.append(["Ledger de tokens del run"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    ledger = logs.get("token_ledger", []) or []
    if ledger:
        ws.append(["Mes", "Movimiento", "Unidades", "Costo unitario",
                   "Precio unitario", "Monto"])
        _style_header(ws, ws.max_row, 6)
        for mv in ledger:
            ws.append([
                _month_of(mv.get("month"), months),
                mv.get("movement_type", ""),
                _num(mv.get("units")),
                _num(mv.get("unit_cost")),
                _num(mv.get("unit_price")),
                _num(mv.get("amount")),
            ])
            row = ws.max_row
            ws.cell(row=row, column=3).number_format = COUNT_FMT
            for col in (4, 5, 6):
                ws.cell(row=row, column=col).number_format = MONEY_FMT
    else:
        ws.append(["Sin movimientos de tokens en este run"])

    # --- Hiring (fase 6; pantalla 49) ---
    ws = wb.create_sheet("Hiring")
    _sheet_run_header(ws, run, "Equipo, nómina y capacidad de onboarding")
    ws.append(["Roles de hiring congelados en el snapshot"])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    roles = snapshot.get("hiring_roles", []) or []
    if roles:
        ws.append(["Rol", "Departamento", "Headcount", "Salario mensual",
                   "Mes de inicio", "Mes de fin", "Ramp (meses)",
                   "Capacidad de onboarding por FTE"])
        _style_header(ws, ws.max_row, 8)
        for r in roles:
            ws.append([
                r.get("name", ""),
                r.get("department", ""),
                r.get("headcount"),
                _num(r.get("monthly_salary")),
                r.get("start_month"),
                r.get("end_month"),
                r.get("ramp_months"),
                _num(r.get("onboarding_capacity_per_fte")),
            ])
            row = ws.max_row
            ws.cell(row=row, column=4).number_format = MONEY_FMT
            ws.cell(row=row, column=8).number_format = COUNT_FMT
    else:
        ws.append(["Sin roles de hiring congelados en este run"])
    ws.append([])
    _append_metric_matrix(ws, months, metrics, HIRING_ROWS)

    file_name = f"pigui_business_plan_{scenario_info['type']}_{run.id[:8]}.xlsx"
    path = os.path.join(EXPORT_DIR, file_name)
    wb.save(path)
    return path
