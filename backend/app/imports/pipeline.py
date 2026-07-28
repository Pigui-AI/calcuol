"""Pipeline de importación asistida (sección 8, IA-01…IA-07).

Funciones puras: bytes → tablas normalizadas → clasificación → mapeo con
confianza → validaciones → propuestas. Nada se persiste aquí; el commit ocurre
solo tras la revisión humana (regla 1.2: "la IA propone y explica; el usuario
confirma antes de persistir").

El extractor es determinista y explicable: reconoce encabezados por sinónimos y
califica cada mapeo. `propose_from_files` acepta un `provider` opcional para
enchufar más adelante un modelo de lenguaje que sugiera mapeos adicionales; las
propuestas de cualquier origen pasan por la misma revisión y validación.
"""
import csv
import hashlib
import io
import re
import unicodedata
import zipfile
from decimal import Decimal, InvalidOperation

from app.engine.money import D, ZERO, ONE

SUPPORTED = ("xlsx", "csv", "pdf", "docx")

# Bandas de confianza de la sección 8.2
CONF_ALTA, CONF_MEDIA = Decimal("0.90"), Decimal("0.70")

# Sinónimos de encabezado por entidad y campo interno (sección 8.3)
FIELD_SYNONYMS = {
    "Client": {
        "trade_name": ["nombre comercial", "nombre del negocio", "negocio", "comercio", "trade name", "marca"],
        "legal_name": ["razon social", "nombre legal", "legal name", "empresa"],
        "industry": ["industria", "giro", "sector", "categoria del negocio", "industry"],
        "currency": ["moneda", "divisa", "currency"],
        "contact_email": ["email", "correo", "correo electronico", "e-mail"],
        "contact_phone": ["telefono", "tel", "celular", "phone"],
        "contact_name": ["contacto", "nombre de contacto", "responsable"],
    },
    "Branch": {
        "name": ["sucursal", "nombre de sucursal", "branch", "punto de venta", "tienda"],
        "location": ["direccion", "ubicacion", "domicilio", "location", "address"],
        "timezone": ["zona horaria", "timezone", "huso"],
        "monthly_capacity": ["capacidad", "capacidad mensual", "capacidad de atencion"],
    },
    "ProductService": {
        "name": ["producto", "servicio", "articulo", "nombre", "item", "descripcion"],
        "sku": ["sku", "codigo", "clave", "codigo de producto"],
        "category": ["categoria", "familia", "linea", "category"],
        "sale_price": ["precio", "precio de venta", "precio unitario", "pvp", "sale price", "venta"],
        "direct_cost": ["costo", "costo directo", "costo unitario", "cost", "compra"],
        "monthly_inventory": ["inventario", "existencias", "stock"],
        "monthly_capacity": ["capacidad mensual", "capacidad"],
        "reward_eligible": ["elegible", "elegible rewards", "aplica recompensa", "reward"],
    },
    "ClientBaseline": {
        "avg_monthly_sales": ["ventas mensuales", "ventas promedio", "ventas", "ingresos mensuales", "facturacion"],
        "avg_monthly_transactions": ["transacciones", "tickets", "operaciones", "ventas por mes (n)", "num transacciones"],
        "avg_ticket": ["ticket promedio", "ticket", "venta promedio"],
        "margin_pct": ["margen", "margen %", "margen porcentaje", "utilidad %"],
        "active_consumers": ["consumidores activos", "clientes activos", "usuarios activos"],
        "registered_consumers": ["consumidores registrados", "clientes registrados", "base de clientes"],
        "monthly_buyers": ["compradores", "compradores mensuales", "clientes que compran"],
        "purchase_frequency": ["frecuencia", "frecuencia de compra", "compras por cliente"],
    },
    "CostItem": {
        "name": ["concepto", "costo", "gasto", "partida", "nombre"],
        "category": ["categoria", "rubro", "area"],
        "behavior": ["comportamiento", "tipo", "behavior"],
        "amount": ["monto", "importe", "cantidad", "valor mensual"],
        "effective_from": ["desde", "mes inicial", "inicio"],
        "effective_to": ["hasta", "mes final", "fin"],
    },
}

# Campos que identifican a la entidad (para nombrar y agrupar filas)
KEY_FIELD = {"Client": "trade_name", "Branch": "name", "ProductService": "name",
             "ClientBaseline": "avg_monthly_sales", "CostItem": "name"}

MONEY_FIELDS = {"sale_price", "direct_cost", "amount", "avg_monthly_sales", "avg_ticket"}
PCT_FIELDS = {"margin_pct"}
INT_FIELDS = {"monthly_inventory", "monthly_capacity", "active_consumers",
              "registered_consumers", "monthly_buyers", "effective_from", "effective_to"}

CURRENCY_TOKENS = {"mxn": "MXN", "usd": "USD", "eur": "EUR", "$": None, "€": "EUR"}


def normalize(text) -> str:
    """Minúsculas sin acentos ni signos: base para comparar encabezados."""
    if text is None:
        return ""
    s = unicodedata.normalize("NFKD", str(text))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9%\s]", " ", s.lower()).strip()


def sha256_of(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def kind_of(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext if ext in SUPPORTED else ""


# --------------------------------------------------------------------------
# IA-02: parseo
# --------------------------------------------------------------------------
def parse_xlsx(data: bytes) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    tables = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        if len(rows) < 2:
            continue
        tables.append({"name": ws.title, "headers": rows[0], "rows": rows[1:]})
    wb.close()
    return {"tables": tables, "text_blocks": []}


def parse_csv(data: bytes) -> dict:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = [r for r in csv.reader(io.StringIO(text), dialect)
            if any(c.strip() for c in r)]
    if len(rows) < 2:
        return {"tables": [], "text_blocks": []}
    return {"tables": [{"name": "csv", "headers": rows[0], "rows": rows[1:]}], "text_blocks": []}


def parse_docx(data: bytes) -> dict:
    """Párrafos y tablas de un .docx con la biblioteca estándar."""
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    paragraphs = []
    for p in re.findall(r"<w:p[ >].*?</w:p>", xml, re.S):
        text = "".join(re.findall(r"<w:t(?: [^>]*)?>(.*?)</w:t>", p, re.S))
        text = re.sub(r"&[a-z]+;", " ", text).strip()
        if text:
            paragraphs.append(text)
    tables = []
    for i, tbl in enumerate(re.findall(r"<w:tbl>.*?</w:tbl>", xml, re.S)):
        rows = []
        for tr in re.findall(r"<w:tr[ >].*?</w:tr>", tbl, re.S):
            cells = ["".join(re.findall(r"<w:t(?: [^>]*)?>(.*?)</w:t>", tc, re.S)).strip()
                     for tc in re.findall(r"<w:tc>.*?</w:tc>", tr, re.S)]
            if any(cells):
                rows.append(cells)
        if len(rows) >= 2:
            tables.append({"name": f"tabla {i + 1}", "headers": rows[0], "rows": rows[1:]})
    return {"tables": tables, "text_blocks": paragraphs}


def parse_pdf(data: bytes) -> dict:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise RuntimeError("Lectura de PDF no disponible en este entorno (falta pypdf)")
    reader = PdfReader(io.BytesIO(data))
    blocks = []
    for page in reader.pages:
        text = (page.extract_text() or "").strip()
        if text:
            blocks.extend([ln.strip() for ln in text.splitlines() if ln.strip()])
    return {"tables": [], "text_blocks": blocks}


def parse_file(filename: str, data: bytes) -> dict:
    """Normaliza cualquier fuente soportada a tablas + bloques de texto."""
    kind = kind_of(filename)
    result = {"kind": kind, "filename": filename, "sha256": sha256_of(data),
              "size_bytes": len(data), "tables": [], "text_blocks": [], "error": None}
    if not kind:
        result["error"] = "Extensión no soportada (usa XLSX, CSV, PDF o DOCX)"
        return result
    try:
        parsed = {"xlsx": parse_xlsx, "csv": parse_csv,
                  "docx": parse_docx, "pdf": parse_pdf}[kind](data)
        result.update(parsed)
    except Exception as exc:  # noqa: BLE001 — el archivo puede venir dañado
        result["error"] = f"No se pudo leer el archivo: {exc}"
    return result


# --------------------------------------------------------------------------
# IA-03: clasificación y mapeo de columnas
# --------------------------------------------------------------------------
def match_header(header, entity_type: str) -> tuple[str | None, Decimal]:
    """Campo interno propuesto para un encabezado y su confianza."""
    norm = normalize(header)
    if not norm:
        return None, ZERO
    best, best_conf = None, ZERO
    for field, synonyms in FIELD_SYNONYMS[entity_type].items():
        for syn in synonyms:
            if norm == syn:
                return field, Decimal("0.95")
            if norm.startswith(syn) or syn in norm:
                conf = Decimal("0.80") if len(syn) > 4 else Decimal("0.72")
                if conf > best_conf:
                    best, best_conf = field, conf
    return best, best_conf


def classify_table(table: dict) -> dict:
    """Entidad más probable para una tabla, con el mapeo de sus columnas (8.1 paso 3)."""
    scored = []
    for entity_type in FIELD_SYNONYMS:
        mapping, confs = {}, []
        for idx, header in enumerate(table["headers"]):
            field, conf = match_header(header, entity_type)
            if field and field not in mapping:
                mapping[field] = {"index": idx, "header": header, "confidence": str(conf)}
                confs.append(conf)
        if not confs:
            continue
        coverage = D(len(mapping)) / D(len(FIELD_SYNONYMS[entity_type]))
        score = (sum(confs, ZERO) / D(len(confs))) * (Decimal("0.5") + Decimal("0.5") * coverage)
        # la entidad necesita su campo clave para poder identificar filas
        if KEY_FIELD[entity_type] not in mapping:
            score *= Decimal("0.4")
        scored.append({"entity_type": entity_type, "mapping": mapping, "score": score})
    if not scored:
        return {"entity_type": None, "mapping": {}, "score": "0"}
    best = max(scored, key=lambda s: s["score"])
    return {"entity_type": best["entity_type"], "mapping": best["mapping"],
            "score": str(best["score"].quantize(Decimal("0.0001")))}


# --------------------------------------------------------------------------
# IA-04: normalización de valores y validaciones
# --------------------------------------------------------------------------
def parse_number(raw) -> tuple[Decimal | None, str | None]:
    """Número y moneda detectada de un valor crudo ('$1,234.50 MXN' → 1234.50, MXN)."""
    if raw is None or str(raw).strip() == "":
        return None, None
    if isinstance(raw, (int, float, Decimal)):
        return D(str(raw)), None
    text = str(raw).strip()
    currency = None
    low = text.lower()
    for token, iso in CURRENCY_TOKENS.items():
        if token in low and iso:
            currency = iso
            break
    cleaned = re.sub(r"[^\d,.\-%]", "", text)
    is_pct = "%" in text
    cleaned = cleaned.replace("%", "")
    if "," in cleaned and "." in cleaned:          # 1,234.56 → 1234.56
        cleaned = cleaned.replace(",", "")
    elif cleaned.count(",") == 1 and len(cleaned.split(",")[-1]) in (1, 2):
        cleaned = cleaned.replace(",", ".")        # 1234,56 → 1234.56
    else:
        cleaned = cleaned.replace(",", "")
    try:
        value = D(cleaned)
    except (InvalidOperation, ValueError):
        return None, currency
    if is_pct:
        value = value / D("100")
    return value, currency


def normalize_value(field: str, raw) -> tuple[str | None, str | None, str | None]:
    """(valor normalizado, moneda, error) según el tipo del campo destino."""
    if field == "reward_eligible":
        return ("true" if normalize(raw) in ("si", "sí", "true", "1", "x", "yes") else "false"), None, None
    if field in MONEY_FIELDS | PCT_FIELDS | INT_FIELDS | {"purchase_frequency",
                                                          "avg_monthly_transactions"}:
        value, currency = parse_number(raw)
        if value is None:
            return None, None, "No se pudo interpretar el número"
        if field in PCT_FIELDS and value > ONE:
            value = value / D("100")     # 35 → 0.35 (16.1: porcentajes como decimal)
        if field in INT_FIELDS:
            value = value.quantize(Decimal("1"))
        if value < ZERO:
            return str(value), currency, "Valor negativo"
        return str(value), currency, None
    text = str(raw).strip()
    return (text or None), None, (None if text else "Vacío")


def validate_entity_rows(entity_type: str, rows: list) -> list:
    """Reglas de la sección 8.1 paso 6 sobre las filas ya normalizadas."""
    issues = []
    seen = {}
    currencies = set()
    for row in rows:
        values = row["values"]
        ref = row["entity_ref"]
        for v in values.values():
            if v.get("currency"):
                currencies.add(v["currency"])
        key = normalize(values.get(KEY_FIELD[entity_type], {}).get("value"))
        if key:
            if key in seen:
                issues.append({"entity_ref": ref, "code": "duplicado", "severity": "media",
                               "message": f"Fila duplicada de «{key}» (también en {seen[key]})"})
            else:
                seen[key] = ref
        if entity_type == "ProductService":
            price = values.get("sale_price", {}).get("value")
            cost = values.get("direct_cost", {}).get("value")
            if price is not None and D(price) <= ZERO:
                issues.append({"entity_ref": ref, "code": "precio_invalido", "severity": "alta",
                               "message": "El precio de venta debe ser mayor a cero"})
            if price is not None and cost is not None and D(cost) > D(price):
                issues.append({"entity_ref": ref, "code": "margen_negativo", "severity": "alta",
                               "message": "El costo directo supera al precio de venta"})
        if entity_type == "ClientBaseline":
            sales = values.get("avg_monthly_sales", {}).get("value")
            tx = values.get("avg_monthly_transactions", {}).get("value")
            ticket = values.get("avg_ticket", {}).get("value")
            if sales and tx and ticket and D(tx) > ZERO:
                implied = D(sales) / D(tx)
                if implied > ZERO and abs(implied - D(ticket)) / implied > D("0.10"):
                    issues.append({"entity_ref": ref, "code": "ticket_inconsistente",
                                   "severity": "media",
                                   "message": f"El ticket declarado ({ticket}) difiere más de 10% "
                                              f"del implícito por ventas/transacciones ({implied:.2f})"})
            active = values.get("active_consumers", {}).get("value")
            registered = values.get("registered_consumers", {}).get("value")
            if active and registered and D(active) > D(registered):
                issues.append({"entity_ref": ref, "code": "activos_mayores",
                               "severity": "alta",
                               "message": "Los consumidores activos superan a los registrados"})
            buyers = values.get("monthly_buyers", {}).get("value")
            if buyers and active and D(buyers) > D(active):
                issues.append({"entity_ref": ref, "code": "compradores_mayores",
                               "severity": "alta",
                               "message": "Los compradores superan a los consumidores activos"})
    if len(currencies) > 1:
        issues.append({"entity_ref": "", "code": "monedas_mezcladas", "severity": "alta",
                       "message": "El archivo mezcla monedas: " + ", ".join(sorted(currencies))})
    return issues


# --------------------------------------------------------------------------
# IA-03/05: propuestas
# --------------------------------------------------------------------------
def propose_from_parsed(parsed: dict, target: str = "auto") -> dict:
    """Propuestas por archivo ya parseado: una por campo reconocido.

    `target` distinto de "auto" fuerza la entidad destino. La confianza combina
    la del encabezado con la calidad del valor; los errores de normalización
    bajan la propuesta a la banda que exige revisión explícita (8.2).
    """
    proposals, issues, tables_info = [], [], []
    for table in parsed.get("tables", []):
        classified = classify_table(table)
        entity_type = classified["entity_type"]
        if target != "auto":
            forced = {"clients": "Client", "catalog": "ProductService",
                      "baseline": "ClientBaseline", "costs": "CostItem"}.get(target)
            if forced:
                classified = {"entity_type": forced,
                              "mapping": classify_table({"headers": table["headers"],
                                                         "rows": table["rows"]})["mapping"],
                              "score": classified["score"]}
                # remapear explícitamente contra la entidad forzada
                mapping = {}
                for idx, header in enumerate(table["headers"]):
                    field, conf = match_header(header, forced)
                    if field and field not in mapping:
                        mapping[field] = {"index": idx, "header": header, "confidence": str(conf)}
                classified["mapping"] = mapping
                entity_type = forced
        if not entity_type or not classified["mapping"]:
            tables_info.append({"table": table["name"], "entity_type": None,
                                "score": classified["score"], "rows": len(table["rows"])})
            continue
        tables_info.append({"table": table["name"], "entity_type": entity_type,
                            "score": classified["score"], "rows": len(table["rows"])})
        normalized_rows = []
        for r_idx, row in enumerate(table["rows"]):
            values = {}
            for field, meta in classified["mapping"].items():
                idx = meta["index"]
                raw = row[idx] if idx < len(row) else None
                value, currency, error = normalize_value(field, raw)
                if value is None and error == "Vacío":
                    continue
                base_conf = D(meta["confidence"])
                conf = base_conf if error is None else min(base_conf, Decimal("0.60"))
                values[field] = {"value": value, "raw": "" if raw is None else str(raw),
                                 "currency": currency, "error": error,
                                 "confidence": str(conf),
                                 "locator": f"{table['name']}!{chr(65 + idx) if idx < 26 else idx}{r_idx + 2}"}
            if not values:
                continue
            key_value = values.get(KEY_FIELD[entity_type], {}).get("value")
            entity_ref = str(key_value or f"{table['name']}:{r_idx + 2}")
            normalized_rows.append({"entity_ref": entity_ref, "values": values})
        issues.extend(validate_entity_rows(entity_type, normalized_rows))
        for row in normalized_rows:
            for field, v in row["values"].items():
                proposals.append({
                    "entity_type": entity_type, "entity_ref": row["entity_ref"],
                    "field_name": field, "locator": v["locator"], "raw_value": v["raw"],
                    "proposed_value": v["value"], "unit": "%" if field in PCT_FIELDS else "",
                    "confidence": v["confidence"],
                    "source_type": "declarado" if v["error"] is None else "estimado",
                    "notes": v["error"] or "",
                })
    return {"proposals": proposals, "issues": issues, "tables": tables_info,
            "text_blocks": len(parsed.get("text_blocks", []))}


def confidence_band(value) -> str:
    """Banda de la sección 8.2 para un valor de confianza."""
    conf = D(value)
    if conf >= CONF_ALTA:
        return "alta"
    if conf >= CONF_MEDIA:
        return "media"
    return "baja"


def summarize(all_proposals: list, all_issues: list) -> dict:
    """Resumen del job: confianza media y conteos por banda y por entidad."""
    if all_proposals:
        avg = sum((D(p["confidence"]) for p in all_proposals), ZERO) / D(len(all_proposals))
    else:
        avg = ZERO
    bands = {"alta": 0, "media": 0, "baja": 0}
    entities: dict[str, int] = {}
    for p in all_proposals:
        bands[confidence_band(p["confidence"])] += 1
        entities[p["entity_type"]] = entities.get(p["entity_type"], 0) + 1
    return {
        "confidence": str(avg.quantize(Decimal("0.0001"))),
        "proposals": len(all_proposals),
        "by_band": bands,
        "by_entity": entities,
        "issues": len(all_issues),
        "issues_by_severity": {
            "alta": sum(1 for i in all_issues if i["severity"] == "alta"),
            "media": sum(1 for i in all_issues if i["severity"] == "media"),
        },
    }
